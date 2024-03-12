import copy
import glob
import os
import re
from logging.config import dictConfig

import openpyxl


class LOG2:
    def __init__(self):
        dictConfig({
            'version': 1,
            'formatters': {
                'default': {
                    'format': '[%(asctime)s] %(message)s',
                }
            },
            'handlers': {
                'file': {
                    'level': 'INFO',
                    'class': 'logging.FileHandler',
                    'filename': './logFile/bizLogic.log',
                    'formatter': 'default',
                    'encoding':'utf-8'
                },
            },
            'root': {
                'level': 'INFO',
                'handlers': ['file']
            }
        })

class ExcepClass(LOG2):
    def __init__(self, fileName):
        self.createDirectory('./logFile') # 로그 셋팅전에 로그폴더 생성해줌
        super().__init__()
        self.file = fileName
        self.wb = openpyxl.load_workbook(self.file)
        # self.hostname = 'hostName'
        # self.upTime = 'Uptime'
        # self.totalNumberOfEntries = 'show ip arp vrf all'

        # 시트 열기(활성화)
        self.sheet = self.wb.active
        (f'excelFilename [{fileName}]')

    def stringSearch(self, fileName, *methods) -> []:
        returnArr = []
        with open(fileName , encoding='UTF-8') as temp_f:
            datafile = temp_f.readlines()
            jsonParam = {}
            for method in methods:
                for line in datafile:
                    jsonParam = method(line , jsonParam)

            returnArr.append(jsonParam)
            (f'hostname uptime info [{jsonParam}]')

        return returnArr

    def createDirectory(self, directory):
        try:
            if not os.path.exists(directory):
                os.makedirs(directory)
        except OSError:
            print("Error: Failed to create the directory.")

    def findLogFile(self) -> []:
        returnArr = []
        all_folder = glob.glob('*') ## 또는 glob.glob('**')
        all_file = [x for x in all_folder if os.path.isfile(x)]
        for file_name in all_file:
            search = ".log"
            if search in file_name:
                returnArr.append(file_name)
        return returnArr

    def getNumber(self, starget , positionUpdownint) -> int:
        number = re.sub(r'[^0-9]' , '',starget)
        return int(number) + positionUpdownint

    # def modifyKernerTime(self,data):
    #     target_value = ''
    #
    #     returnArr = []
    #     jsonParam = {}
    #
    #     matchingHostName = {}
    #     for i in data[0]:
    #         target_value = i
    #         excelColumnPosition = data[0][i]['excelColumnPosition']
    #         excelRowPosition = data[0][i]['excelRowPosition']
    #         searchExcelToString = data[0][i]['searchExcelToString']
    #         changeData = data[0][i]['data']
    #
    #         if target_value in 'hostname': continue
    #
    #
    #         for row in self.sheet.iter_rows():
    #             # jsonParam = {'Hostname':cell.value ,'excelColumnPosition':excelColumnPosition, 'excelRowPosition':excelRowPosition};
    #             for cell in row:
    #                 if cell.value==searchExcelToString:
    #                     logNamePositionNumber = self.getNumber(cell.coordinate , 0)
    #
    #                 if cell.value == data[0]['hostname']['data']:
    #                     matchingHostName = {'hostname':cell.value}
    #                     logNamePositionNumber = self.getNumber(cell.coordinate , 0)
    #                     logNamePositionNumber = self.getNumber(cell.coordinate , excelRowPosition)
    #                     self.sheet.cell(row=logNamePositionNumber, column=excelColumnPosition).value = changeData
    #                     # print(f'self.sheet.cell(row=logNamePositionNumber, column=excelColumnPosition).value :: {self.sheet.cell(row=logNamePositionNumber, column=excelColumnPosition).value}')
    #                     (f'targetDat'
    #                      f'a[{target_value}] excel modify position row[{logNamePositionNumber}] column[{excelColumnPosition}] > change [{changeData}]')
    #
    #                 if matchingHostName == data[0]['hostname']['data'] and cell.value==searchExcelToString:
    #                     matchingHostName = {searchExcelToString:cell.value}


    def getExcelDataReset(self, fileDatas) -> {}:
        exceldata = {}
        # 'Uptime': {'data': '2111 day(s)', 'searchExcelToString': 'Uptime', 'excelColumnPosition': 24},
        for key in fileDatas[0][0]:
            exceldata[key] = {'data':'', 'excelColumnPosition':fileDatas[0][0][key]['excelColumnPosition'] }
        return exceldata

    def excelDataAppend(self ,exceldata , excelDataArr ,origenData) -> {}:
        for key in exceldata:
            if exceldata[key]['data'] == '':
                return {'exceldata':exceldata ,'excelDataArr':excelDataArr }

        excelDataArr.append(exceldata)
        tmpExceldata = copy.deepcopy(origenData)
        # for key in exceldata:
        #     print(key)
        #     print(exceldata)
        #     tmpExceldata[key]['data'] = ''

        return {'exceldata':tmpExceldata ,'excelDataArr':excelDataArr }

    def getExcelPostionData(self , fileDatas) -> []:
        # exceldata = {'Hostname':'','show mac address-table':'', 'show ip arp vrf all':'' , 'Uptime':''}
        exceldata = self.getExcelDataReset(fileDatas)
        origenData = copy.deepcopy(exceldata)
        print(f'exceldata초기화 : [{exceldata}]')
        print(f'fileDatas : [{fileDatas}]')

        excelDataArr = []
        for row in self.sheet.iter_rows():

            # if exceldata['Hostname'] != '' and  exceldata['show mac address-table'] != '' and exceldata['show ip arp vrf all'] != '' and exceldata['Uptime'] != '':
            #     excelDataArr.append(exceldata)
            #     exceldata = {'Hostname':'','show mac address-table':'', 'show ip arp vrf all':'' , 'Uptime':''}


            jsonReturnData = self.excelDataAppend(exceldata , excelDataArr , origenData)
            exceldata = jsonReturnData['exceldata']
            excelDataArr = jsonReturnData['excelDataArr']

            for cell in row:
                for key in origenData:
                    if cell.value==key:
                        logNamePositionNumber = self.getNumber(cell.coordinate , 0)
                        exceldata[key] = {'data':self.sheet.cell(row=logNamePositionNumber, column=origenData[key]['excelColumnPosition']).value , 'row':logNamePositionNumber ,'column':origenData[key]['excelColumnPosition'] }
                        continue

                    # if cell.value=='show mac address-table' and '' == exceldata['show mac address-table']:
                    #     logNamePositionNumber = self.getNumber(cell.coordinate , 0)
                    #     exceldata['show mac address-table'] = {'row':logNamePositionNumber ,'column':26}
                    #     continue
                    # if cell.value=='show ip arp vrf all' and '' == exceldata['show ip arp vrf all']:
                    #     logNamePositionNumber = self.getNumber(cell.coordinate , 0)
                    #     exceldata['show ip arp vrf all'] = {'row':logNamePositionNumber ,'column':26}
                    #     continue
                    #
                    # if cell.value=='Uptime' and '' == exceldata['Uptime']:
                    #     logNamePositionNumber = self.getNumber(cell.coordinate , 0)
                    #     exceldata['Uptime'] = {'row':logNamePositionNumber ,'column':24}
                    #     continue
        return excelDataArr

    # def exec(self, data):
    #     self.modifyKernerTime(data)

    def secondExec(self , fileDatas , excelLoopArrayData):
        for excelData in excelLoopArrayData:
            for filedata in fileDatas:
                if filedata[0]['Hostname']['data'] == excelData['Hostname']['data']:
                    # print(filedata[0])
                    # print(excelData)
                    self.addSheetCell(excelData , filedata[0])


    def addSheetCell(self, excelData , filedata):
        for key in excelData:
            try:
                self.sheet.cell(row=excelData[key]['row'], column=excelData[key]['column']).value = filedata[key]['data']
            except:
                try:
                    if 'show mac address-table' == key:
                        self.sheet.cell(row=excelData[key]['row'], column=excelData[key]['column']).value = filedata['show mac address-table_local']['data']
                except:pass


            # try:
            #     self.sheet.cell(row=excelData['show mac address-table']['row'], column=excelData['show mac address-table']['column']).value = filedata['show mac address-table']['data']
            # except:
            #     try:
            #         self.sheet.cell(row=excelData['show mac address-table']['row'], column=excelData['show mac address-table']['column']).value = filedata['show mac address-table_local']['data']
            #     except:pass
            #
            # try:
            #     self.sheet.cell(row=excelData['Uptime']['row'], column=excelData['Uptime']['column']).value = filedata['Uptime']['data']
            # except:pass
            #
            #
            # try:
            #     self.sheet.cell(row=excelData['show ip arp vrf all']['row'], column=excelData['show ip arp vrf all']['column']).value = filedata['show ip arp vrf all']['data']
            # except:pass

    def execute(self):
        fileDatas = [self.stringSearch(i
                                       , self.searchHostName
                                       , self.kernelUptimeSet
                                       , self.totalNumberOfEntriesSet
                                       , self.dynamicAddressCountSet
                                       , self.dynamicAddressCountLocalSet
                                       ) for i in self.findLogFile()]

        excelLoopPostionArr = self.getExcelPostionData(fileDatas)

        # print(f'excelLoopPostionArr :: [{excelLoopPostionArr}]')

        # 엑셀의 파일명과 바꿔야될 포지션 정보를 모두 담았다.
        self.secondExec(fileDatas, excelLoopPostionArr)

        # [self.exec(i) for i in fileDatas]
        self.wb.save(self.file)


    def searchHostName(self, line , jsonParam) -> any:
        searchFileToString = 'hostname'
        excelColumnPosition = 14;
        if searchFileToString in line:
            tmpData = line.split(searchFileToString)[1]
            tmpData = tmpData.replace("\n", "").strip()
            jsonParam['Hostname'] = {'data':tmpData,'searchExcelToString':searchFileToString, 'excelRowPosition':-1 , 'excelColumnPosition':excelColumnPosition};
        return jsonParam;

    def kernelUptimeSet(self, line , jsonParam) -> any:
        """ 포지션 기준은 hostname 기준으로  """
        searchFileToString = 'Kernel uptime is'
        searchExcelToString = 'Uptime'
        # excelRowPosition = 1
        excelColumnPosition = 24;
        if searchFileToString in line:
             tmpData = line.split(searchFileToString)[1]
             tmpData = tmpData.split(',')[0]
             tmpData = tmpData.replace("\n", "").strip()
             jsonParam[searchExcelToString] = {'data':tmpData,'searchExcelToString':searchExcelToString , 'excelColumnPosition':excelColumnPosition};
             # jsonParam[searchExcelToString] = {'data':tmpData,'searchExcelToString':searchExcelToString,'excelColumnPosition':excelColumnPosition, 'excelRowPosition':excelRowPosition};
        return jsonParam;

    def totalNumberOfEntriesSet(self, line , jsonParam) -> any:
        """ 포지션 기준은 hostname 기준으로  """
        searchFileToString = 'Total number of entries'
        searchExcelToString = 'show ip arp vrf all'
        # excelRowPosition = 51
        excelColumnPosition = 26;
        if searchFileToString in line:
            tmpData = line.split(':')[1]
            tmpData = tmpData.replace("\n", "").strip()
            jsonParam[searchExcelToString] = {'data':tmpData+' (개)','searchExcelToString':searchExcelToString , 'excelColumnPosition':excelColumnPosition}
            # jsonParam[searchExcelToString] = {'data':tmpData+' (개)','searchExcelToString':searchExcelToString,'excelColumnPosition':excelColumnPosition, 'excelRowPosition':excelRowPosition};
        return jsonParam;

    def dynamicAddressCountSet(self, line , jsonParam) -> any:
        """ 포지션 기준은 hostname 기준으로  """
        searchFileToString = 'Dynamic Address Count'
        searchExcelToString = 'show mac address-table'
        # excelRowPosition = 50
        excelColumnPosition = 26;
        if searchFileToString in line:
            try:
                tmpData = line.split(':')[1]
                tmpData = tmpData.replace("\n", "").strip()
                jsonParam[searchExcelToString] = {'data':tmpData+' (개)','searchExcelToString':searchExcelToString, 'excelColumnPosition':excelColumnPosition};
                # jsonParam[searchExcelToString] = {'data':tmpData+' (개)','searchExcelToString':searchExcelToString,'excelColumnPosition':excelColumnPosition, 'excelRowPosition':excelRowPosition};
            except:
                print('except')
        return jsonParam;

    def dynamicAddressCountLocalSet(self, line , jsonParam) -> any:
        """ 포지션 기준은 hostname 기준으로  """
        searchFileToString = 'Dynamic Local Address Count'
        searchExcelToString = 'show mac address-table_local'
        # excelRowPosition = 50
        excelColumnPosition = 26;
        if searchFileToString in line:
            try:
                tmpData = line.split(':')[1]
                tmpData = tmpData.replace("\n", "").strip()
                jsonParam[searchExcelToString] = {'data':tmpData+' (개)','searchExcelToString':searchExcelToString ,'excelColumnPosition':excelColumnPosition};
                # jsonParam[searchExcelToString] = {'data':tmpData+' (개)','searchExcelToString':searchExcelToString,'excelColumnPosition':excelColumnPosition, 'excelRowPosition':excelRowPosition};
            except:
                print('except')
        return jsonParam;


if __name__ == '__main__':
    fileName = ''
    print(f'같은 폴더내 .log 파일 및 엑셀 파일이 같이 있어야됨 \n')
    print(f'엑셀파일명지정해줘야됨 정상파일이 지정되면 실행함 : ex) sample.xlsx \n')
    print(f'exit() : 실행 안하고 그냥 나가기 \r\n')
    while True:
        message = input('')
        try:
            if message in 'exit()':
                exit(0)
                break;

            fileName = message
            print(f"엑셀 파일명 : {fileName}")
            break;
        except:
            print('정상정신 형식으로 엑셀파일을 지정해주세요')

    print(f'엑셀파일명: {fileName} 으로 실행!!')
    excelC = ExcepClass(fileName.strip())
    excelC.execute();